import os
import sys
import shutil
import string
import pprint
import platform
import warnings
import openpyxl
from openpyxl.styles import Font, Alignment
#from send2trash import send2trash
warnings.filterwarnings("ignore")

#TODO: Gjør mulig med "Vil gjøre dette_____", få de til å skrive det under kommentarer?

deltakere = []
deltakskrav = {}
#VERKTØY


# Importerer malen som ordboks
def importer_mal():
    global mal, c_a, c_c
    mal = {}
    wb = openpyxl.load_workbook('Malen.xlsx') #workbook
    ws = wb.active # eller: wb.get_sheet_by_name() ##worksheet
    i = 2 #pga malformat
    while True:
        try:
            c_a = ws['A' + str(i)].value
            c_c = ws['C' + str(i)].value
            if c_a == None:
                pass
            elif 'slutt' in c_a.lower():
                print('Mal importert.')
                break #stopper hvis cellen er 'slutt'
            elif '#' in c_a:
                header = fjern_hash(c_a)
                #header = c_a
                mal[header] = []
                lag_deltakskravliste(c_c, c_a)
            elif c_a:
                mal[header].insert(-1, c_a)
            i += 1
        except TypeError as err:
            print('Feil oppstod: {}'.format(err))
            sys.exit()
    #print(max(mal.keys(), key=len))
    #sys.exit()

# Finn ut hvilken gruppe et valg tilhører (Eks: Tog tilhører transport)
def finn_gruppe(valg):
    global mal
    for gruppe, valgliste in mal.items():
        if valg in valgliste:
            return gruppe

def få_bokstav(ting):
    global mal
    for key, value in mal.items():
        if ting in value:
            index_ = mal[key].index(ting)
            letter = string.ascii_uppercase[index_]
            return string.ascii_uppercase[index_]

def fjern_hash(gruppe):
    if '#' in gruppe:
        return gruppe[1:]
    else:
        return gruppe

def gjør_oppfylt_deltakskrav(name, inndata):
    gruppe = finn_gruppe(inndata)
    if gruppe in deltakskrav.keys():
        if name not in oppfylt_deltakskrav[gruppe]:
            oppfylt_deltakskrav[gruppe].append(name)

def lag_deltakskravliste(kravdata, gruppe):
    global deltakskrav, c_a, c_c
    if kravdata:
        deltakskrav[fjern_hash(gruppe)] = True

'''def isdeltakskrav(name, gruppe):
    if gruppe in deltakskrav.keys():
        oppfylt_deltakskrav[gruppe].insert(0, name)
        print(name, oppfylt_deltakskrav[gruppe])
        sys.exit()'''

def new_run():
##    global output_files
    output_files = [
                'Lederønsker.xlsx',
                'Unnasluntrere.txt',
                'Kommentarer.txt',
                'Deltakere.txt']
    dirlist = os.listdir('.')
    for file in output_files:
        if file in dirlist:
            send2trash(file)


# INPUT/OUTPUT
def lag_tomme_lister2():
    #global fontObj, styleObj
    print('Lager tomme lister.')
    global mal
    wb = openpyxl.Workbook()
    ws = wb.active
    #fontObj = Font(name='Times New Roman', bold=True, size=18)
    #styleObj = Style(font=fontObj)
    for gruppe, valgliste in sorted(mal.items(), key=str):
        wb.create_sheet(title=gruppe)
        ws = wb.get_sheet_by_name(gruppe)
        for valg in valgliste:
        #for valg in sorted(valgliste):
            letter = få_bokstav(valg)
            ws[letter + '2'] = valg
        ws.merge_cells('A1:' + letter + '1')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'] = gruppe
        #style_ranges(ws, 'A1:' + letter + '1', font=fontObj, alignment=al)
        #ws['A1'].font = fontObj
        wb.save('Lederønsker.xlsx')
    print('Tomme lister laget.')


def ledernavn_til_liste(navn, valg, dobbel=False):
    global checksum_jobs_registered_in_xlsx
    i = 3
    wb = openpyxl.load_workbook('Lederønsker.xlsx')
    #print('Valg: {}, Gruppe: {}'.format(valg, finn_gruppe(valg))) ##Feilsøk
    ws = wb.get_sheet_by_name(finn_gruppe(valg))
    letter = få_bokstav(valg)

    #print("finngruppe(valg), valg", finn_gruppe(valg), valg)
    #print("bokstav", få_bokstav(valg))



    #Skriver navn på første ledige celle i kolonne hvor valg er gruppemedlem
    #I riktig arbeidsark
    while True:
        if ws[letter + str(i)].value:
            i += 1
        else:
            if dobbel==True:
                pass
                #fontObj = Font(name='Times New Roman', bold=True)
                #styleObj = Style(font=fontObj)
                #ws[letter + str(i)].style = styleObj
            ws[letter + str(i)] = navn
            checksum_jobs_registered_in_xlsx += 1
            #print("bostav og navn = ", ws[letter + str(i)].value)
            wb.save('Lederønsker.xlsx')
            #exit()
            break

def kommentar(name, comment):
    global new_run
    if new_run == True:
        open('Kommentarer.txt', 'w').write('')
    with open('Kommentarer.txt', 'a') as f:
        f.write(name + ':\n' + comment + '\n\n')

def check_participancy():
    global unnasluntrere
    print('Begynner å sjekke for deltakelse.')
    superfirst = True
    for name in deltakere:
        first = True
        with open('Unnasluntrere.txt', 'a') as f:
            for gruppe, name_oppfylt in oppfylt_deltakskrav.items():
                if not name in name_oppfylt:
                    if superfirst:
                        f.write('Her ser du hvem som ikke har meldt seg på påkrevd hva.\n\n')
                        f.write('{}:'.format(name))
                    elif first:
                        f.write('\n\n{}:'.format(name))
                    f.write('\n{}'.format(gruppe))
                    first = False
                    superfirst = False
                #open('Unnasluntrer.txt', 'a').write('{}: {}\n'.format(name,key))
                #print('Unnasluntrer funnet: {}: Mangler {}'.format(name, key))
    print('Ferdig med å sjekke for deltakelse.')

def gjør_jobben_til_susanne():
    global new_run, oppfylt_deltakskrav, checksum_jobs_wanted_by_leaders
    checksum_jobs_wanted_by_leaders = 0
    print('Gjør jobben til Susanne.')
    new_run = True
    oppfylt_deltakskrav = {i:[] for i in deltakskrav.keys()}
    #all_names =[i for i in os.listdir('.') if '##' in i]
    #for index, file_ in enumerature(sorted(all_names)):
    for f in os.listdir('.'):
        if '##' in f:
            wb = openpyxl.load_workbook(f)
            ws = wb.active
            name = f[2:f.index('.')]
            print('Begynner med:', name)
            assert '#' not in name and '.' not in name, '# eller . i navn'
            i = 2
            while True:
                c_a = ws['A' + str(i)]
                c_b = ws['B' + str(i)]
                c_c = ws['C' + str(i)]
                if c_b.value == None:
                    pass
                elif 'slutt' in c_b.value.lower():
                    print('Ferdig med jobben til Susanne.')
                    break
                elif c_b.value is None:
                    pass
                elif 'x' == c_b.value.lower():
                    ledernavn_til_liste(name, fjern_hash(c_a.value))
                    gjør_oppfylt_deltakskrav(name, c_a.value)
                    checksum_jobs_wanted_by_leaders += 1
                elif 'xx' == c_b.value.lower():
                    ledernavn_til_liste(name, fjern_hash(c_a.value), dobbel=True)
                    gjør_oppfylt_deltakskrav(name, c_a.value)
                    checksum_jobs_wanted_by_leaders += 1
                elif c_b.value.lower():
                    kommentar(name, c_b.value)
                i += 1
            deltakere.append(name)

            new_run = False
            #if index % 10 == =:
                #print('Gjenstår {} konfirmanter.'.format(len(all_names) - index))
            print('Ferdig med: {}'.format(name))
    #func check participancy for all participants
    print('Jobben til Susanne gjort.')

def deltakere_til_txt():
    print('Skriver deltakere til Deltakere.txt.')
    with open('Deltakere.txt', 'a') as f:
        f.write('Antall deltakere: {}.\n\n'.format(len(deltakere)))
        for name in sorted(deltakere):
            f.write(name + '\n')
    print('Deltakere skrevet til Deltakere.txt')


def gjør_justeringer():
    #global styleObj, fontObj
    print('Gjør små justeringer.')
    wb = openpyxl.load_workbook('Lederønsker.xlsx')
    for sheet in wb.get_sheet_names():
        ws = wb.get_sheet_by_name(sheet)
        if ws['A1']:
            pass
            #ws['A1'].style = styleObj
            wb.save('Lederønsker.xlsx')

def move(f):
##    print(arbeidsmappe + f, output_mappe + "/" + f)
    try:
        os.rename(os.path.abspath(f), output_mappe + "/" + f)
    except FileNotFoundError:
        print("File not found :" + f)
        
output_files = [
                'Lederønsker.xlsx',
                'Unnasluntrere.txt',
                'Kommentarer.txt',
                'Deltakere.txt']


##name = input()
checksum_jobs_registered_in_xlsx = 0
#navnformat = input('Hva er navn-postfixen? Eks: Knattholmen 2016')
arbeidsmappe = os.path.split(__file__)[0] + '/'
output_mappe = arbeidsmappe + 'Output'
os.makedirs(output_mappe, exist_ok=True)
os.chdir(arbeidsmappe)
#new_run()
importer_mal()
lag_tomme_lister2()
gjør_jobben_til_susanne()
deltakere_til_txt()
check_participancy()
print('Flytter filer til mappe: Output')
for f in output_files:
    move(f)

if checksum_jobs_registered_in_xlsx == checksum_jobs_wanted_by_leaders:
    print('FERDIII >(^^)> <(^^)<')
else:
    print("Fatal error.")

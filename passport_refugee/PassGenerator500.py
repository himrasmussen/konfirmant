import os
import sys
import random
import pprint
import shutil
try:
    import openpyxl
except ImportError:
    import pip
    pip.main(['install', "openpyxl"]) #because we need openpyxl
import warnings
warnings.filterwarnings("ignore")

'''
Glory to Knattholmen! #Arstoztka
'''


genders = ["Boy", "Girl"]

def make_bday():
    '''
    Because a func is cleaner than doing this elswhere.
    '''
    zero_digits = ["0" + str(i) for i in range(10)]
    bday = []
    bday.append(str(random.choice(zero_digits[1:] + list(range(10, 29)))))
    bday.append(str(random.choice(zero_digits[1:] + list(range(10, 13)))))
    bday.append(str(random.choice(list(range(96, 100)) + zero_digits[:5])))
    return bday

def make_prs_num(bday, gender):
    '''
    The Norwegian algorythm for creating unique personal numbers (ID)
    http://www.matematikk.org/artikkel.html?tid=64296
    '''

    prs_num = ''
    if 96 <= int(bday[-1]) <= 99:
        prs_num += str(random.randint(0, 4)) + str(random.randint(0, 9)) #00-50 exclusive
    else:
        prs_num += str(random.randint(50, 100))
    if gender == "Girl":
        prs_num += random.choice(["2", "4", "6", "8"])
    elif gender == "Boy":
        prs_num += random.choice(["1", "3", "5", "7", "9"])

    #pardon the mess
    formula_controldigit_1 = [3, 7, 6, 1, 8, 9, 4, 5, 2]
    #V1=3D1+7D2+6M1+M2+8å1+9å2+4I1+5I2+2I3 #Algorythm for first control digit
    formula_controldigit_2 = [5, 4, 3, 2, 7, 6, 5, 4, 3, 2]
    #V2=5D1+4D2+3M1+2M2+7å1+6å2+5I1+4I2+3I3+2K1 #Algorythm for second control digit

    temp_control_digit = 0
    for index, i in enumerate(''.join(bday) + prs_num):
        temp_control_digit += (int(i) * formula_controldigit_1[index])
        #print(index)
    if temp_control_digit % 11 == 0:
        prs_num += str(0)
    else:
        prs_num += str(11 - (temp_control_digit % 11))

    temp_control_digit = 0
    for index, i in enumerate(''.join(bday) + prs_num):
        temp_control_digit += (int(i) * formula_controldigit_2[index])
        #print(index)
    #prs_num += str(temp_control_digit % 11)
    if temp_control_digit % 11 == 0:
        prs_num += str(0)
    else:
        prs_num += str(11 - (temp_control_digit % 11))

    if len(prs_num) == 6:
        prs_num = prs_num[:5]
    return prs_num



def make_passport(country, gender, bday, i):
    global country_data

    gender_letter = {"Boy": "M", "Girl": "F"}
   # country_data = country_data[country]

    wb = openpyxl.load_workbook("Pass_mal.xlsx")
    #wb = openpyxl.load_workbook("Pass_mal.xlsx")
    ws = wb.get_active_sheet()

    ws["B4"] = country_data[country]["Norwegian name for country"][0].upper()
    ws["C4"] = country_data[country]["English name for country"][0].upper()
    ws["B7"] = random.choice(country_data[country]["Surnames"]).upper()
    ws["D7"] = random.choice(country_data[country][gender + " names"]).upper()
    ws["B10"] = '.'.join(bday)
    ws["D10"] = "{}, {}".format(    random.choice(country_data[country]["Birthplaces"]).upper(),
                                    country_data[country]["Abbreviation"][0].upper()) #.upper
    ws["B13"] = country_data[country]["Norwegian name for nationality"][0].upper() + ' ' + country_data[country]["English name for nationality"][0].upper()
    while True:
        try:
            ws["D13"] = "{} {}".format(
                                        ''.join(bday),
                                        make_prs_num(bday, gender))
            break
        except:
            pass

    ws["A13"] = gender_letter[gender].upper()
    #assert country is str, "country not str"
    #assert gender is str, "gender not str"
    #assert str(i) is str, "str(i) not str"
    path = os.path.join("passports", "##{}-{}-{}.xlsx".format(str(country), str(gender), str(i)))
    wb.save(path)

# because we need the fucking country data
def import_all_country_data():
    global country_data
    for country_data_txt in os.listdir('.'):
        if country_data_txt.endswith('.txt'):
            raw_data = open(country_data_txt).read().splitlines()
            country = country_data_txt[:-4]
            country_data[country] = {}
            for line in raw_data:
                if line:
                    key = line.split(':')[0]
                    values = line.split(':')[1].split(',')
                    country_data[country][key] = values

country_data = {
                "Country1" :
                        {
                        "Norwegian name for country": "Land",
                        "English name for country": "Country",
                        "Abbreviation": "LAN",
                        "Norwegian name for nationality": "Landsk",
                        "English name for nationality": "Country-ian/an",
                        "Boy names": ["Boy", "Boy", "Boy"],
                        "Girl names": ["Girl", "Girl", "Girl"],
                        "Surnames": ["Surname", "Surname", "Surname"],
                        "Birtplaces": ["Birthplace", "Birthplace", "Birthplace"]
                        }
}


country_data = {}
import_all_country_data()
def main():
    global country_data
    shutil.rmtree("passports") if os.path.isdir("passports") else None
    os.makedirs("passports")
    country_list = [country for country in country_data.keys()]
    for country in country_list:
        for gender in ["Boy", "Girl"]:
            range_num = int(input("How many {} {}-passports?".format(country_data[country]["English name for nationality"][0], gender.lower())))
            for i in range(range_num):
                make_passport(country, gender, make_bday(), i)
        wb = openpyxl.Workbook()
        ws = wb.get_active_sheet
        #ws["A1"] = "{}".format(country)
        #openpyxl.Workbook().save("##{}-.xlsx".format(country))

main()
print("DONE")
